from typing import Collection
import requests
import json
import csv
import openpyxl
from openpyxl.styles import PatternFill, Font
from os.path import exists
from datetime import date
from openpyxl.chart import LineChart,Reference
fileExist = True


fileName = 'collections' + '.xlsx'
if(not exists(fileName)):
    print('File',fileName,'not Found, Creating now...')
    workbook = openpyxl.Workbook()
    workbook.save(filename=fileName)
else:
    print('File',fileName,'found...')
def calcPercentIncrease(num1, num2):

    temp = num1 - num2
    return temp / num1
#TODO Figure out how to calculate the colors and convert percent Change to color code
def colorCell(c1,c2):
    red = PatternFill(fill_type='solid',start_color='00FF0000', end_color='FF000000')
    green = PatternFill(fill_type='solid',start_color='0000FF00', end_color='FF000000')
    try:
        if(c1.value > c2.value):
            c1.fill = green
        elif(c1.value < c2.value):
            c1.fill = red
        
    except:
        print('FirstCell in sheet so no color OR no change')
def calcHexColorVal(num):
    return 0
wb = openpyxl.load_workbook('collections.xlsx', read_only=False)   # open an Excel file and return a workbook
def writeRow(row, sheet,rowNum,offset=0):
    alphabet = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S']
    for i in range(len(row)):
        sheet[alphabet[i + offset] + str(rowNum)] = row[i]
    colorCell(sheet['B2'],sheet['B3'])    
    
        
def getFilePath():
    f = open('settings.json')
    data = json.load(f)
    print('CSV File location:',data['path_to_csv_files'],'\n\n')
    return data['path_to_csv_files']
#filePath = getFilePath()
def writeOverview():
    wb['Sheet'].delete_rows(1)
    wb['Sheet'].delete_rows(1)
    wb['Sheet'].insert_rows(1)
    wb['Sheet'].insert_rows(1)
    sheetNames = wb.sheetnames
    sheetNames = sheetNames[1:]
    writeRow(sheetNames,wb['Sheet'],"1", offset=1)
    values = ['Percent Change']
    for sheet in sheetNames:
        if sheet == wb['Sheet']:
            continue
        values.append(calcPercentIncrease(wb[sheet]['B2'].value,wb[sheet]['B3'].value))
    writeRow(values,wb['Sheet'],2)
def collectStats(collectionToFind,cryptoPrice):
    url = "https://api.opensea.io/api/v1/collection/"
    url += collectionToFind 
    url += '/stats'
    
    headers = {"Accept": "application/json"}
    response = requests.request("GET", url, headers=headers)
    list = json.loads(response.text)
    
    if(list['stats']['floor_price']):
        print('\nData recieved for:', collectionToFind)
    else:
        raise Exception("Invalid Collection",collectionToFind, 'Ensure that the name in collecitons.txt is the name that appears in the url')
        
    values = []
    stats = list['stats']
    values.append(date.today().strftime("%m/%d/%Y"))
    values.append(stats['floor_price'])
    values.append(float(values[1]) * float(cryptoPrice))
    values.append(stats['one_day_volume'])
    values.append(stats['one_day_change'])
    values.append(stats['one_day_sales'])

    values.append(stats['seven_day_volume'])
    values.append(stats['seven_day_change'])
    values.append(stats['seven_day_sales'])

    values.append(stats['thirty_day_volume'])
    values.append(stats['thirty_day_change'])
    values.append(stats['thirty_day_sales'])

    values.append(stats['total_volume'])
    values.append(stats['total_sales'])
    values.append(stats['total_supply'])
    values.append(stats['num_owners'])
    values.append(stats['average_price'])
    values.append(float(stats['average_price']) * float(cryptoPrice))
    values.append(stats['market_cap'])
    
    
    return values
def getCryptoExchange(cryptoToFind):
    
    url = 'https://api.coinbase.com/v2/prices/'
    url += cryptoToFind
    url += '-USD/buy'
    crypto_response = requests.request("GET",url)

    CRYPTO_DATA = json.loads(crypto_response.text)
    
    CRYPTO_PRICE = CRYPTO_DATA['data']['amount']
    print('Crypto Price recieved')
    return CRYPTO_PRICE
#TODO figure out how the fuck to get graphs to work
def drawGraph(sheet):
    c1 = LineChart()
    c1.title = 'ETH Price VS. Date'
    c1.style = '13'
    c1.y_axis.title = 'ETH Price'
    c1.x_axis.title = 'Date'
    print(sheet.rows)
    data = Reference(sheet,min_col = 2, min_row = 1, max_col = 2 ,max_row = sheet.max_row)
    c1.add_data(data,titles_from_data = True)
def toCSV(name, row):
    header = ['Date','Floor Price ETH','Floor Price USD','One Day Volume','One Day Change', 'One Day Sales', 'Seven Day Volume','Seven Day Change', 'Seven Day Sales',
                   'Thirty Day Volume','Thirty Day Change', 'Thirty Day Sales','Total Volume', 'Total Sales','Total Supply','Num Owners', 'Average Price ETH', 'Average Price USD','Market Cap']
   
    
    if name in wb.sheetnames:
        print('Sheet',name,'found, adding row to it')
        sheet = wb[name]
        sheet.insert_rows(2)
        writeRow(row,wb[name],"2")
        
    else:
        print('Sheet',name,'NOT found, creating now...')
        wb.create_sheet(name)
        print('Writing to sheet',name)
        wb[name].insert_rows(1)
        writeRow(header,wb[name],"1")
        wb[name].insert_rows(2)
        writeRow(row,wb[name],"2")
    
    writeOverview()
    wb.save('collections.xlsx')
    

    
with open('collections.txt') as collections:
    lines = collections.readlines()
    cryptoPrice = getCryptoExchange('ETH')
    for i in range(len(lines) - 1):
        temp = lines[i][:-1]
        lines[i] = temp
    for line in lines:
        
        stats = (collectStats(line,cryptoPrice))
            
        
        try:
            stats = (collectStats(line,cryptoPrice))
            try:
                toCSV(line, stats)
            except:
                print("\n\n================ERROR WITH CSV FILE================")
                print('Make sure that the .csv file is not open by another program')
        except:
            print('\n\n================INVALID COLLECTION NAME================')
            print('Invalid collection:',line)
            print('Ensure that the name is the one that appears in the url of collection\n\n')
        
        